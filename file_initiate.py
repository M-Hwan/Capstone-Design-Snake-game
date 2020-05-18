# 이주연
# 회원 관리 파일 생성

from openpyxl import Workbook
from openpyxl.styles import Alignment # 셀 값 가운데 정렬 시, 필요모듈

def file_initiate():
    table_attribute = ['','카카오톡 ID','점수','순위']
    table_attribute2 = ['','카카오톡 ID','점수']
    
    wb1 = Workbook()
    wb2 = Workbook()
    ws1 = wb1.active #sheet 활성화
    ws2 = wb2.active #sheet 활성화
    ws1.title = '회원관리' #시트 ws1의 이름을 지정
    ws2.title = '플레이 로그'
    
    for c in range(1,4): 
        ws1.cell(row=1, column = c, value = table_attribute[c]) #릴레이션의 속성 지정
        ws1.cell(row=1, column = c).alignment = Alignment(horizontal = 'center') # 셀값 가운데 정렬

    for c in range(1,3):
        ws2.cell(row=1, column = c, value = table_attribute2[c]) #릴레이션의 속성 지정
        ws2.cell(row=1, column = c).alignment = Alignment(horizontal = 'center')
        
    ws1.column_dimensions['A'].width = 20 # A열의 너비를 20으로 설정
    ws2.column_dimensions['A'].width = 20 # A열의 너비를 20으로 설정
    wb1.save('회원관리.xlsx')
    wb2.save('플레이로그.xlsx')

file_initiate()
