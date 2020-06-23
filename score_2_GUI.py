# -*- coding: utf-8 -*-

# Score implementation generated from reading ui file 'score.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!


# 이주연
# 게임 종료 후, 점수판 GUI(버튼 추가)

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QCursor, QIcon, QColor

class Ui_Score(object):

    def OK_bttnClicked(self):
        #sys.exit()
        Score.hide()
        return
        
    def Restart_bttnClicked(self):
        #Score.hide() # 점수판 화면 숨기기
        import snake
        print('뱀 게임 재시작')
            
    def setupUi(self, Score):

        # 윈도우 아이콘 설정
        Score.setWindowIcon(QIcon('images\game_logo.png'))
        Score.setObjectName("Score")
        Score.resize(560, 740)    ##### 창 크기:(가로,세로)
        self.tableWidget = QtWidgets.QTableWidget(Score)
        # 위젯 위치 조절(가로로 떨어진 거리, 세로로 떨어진 거리, 가로, 세로)
        self.tableWidget.setGeometry(QtCore.QRect(25, 20, 510, 613)) 
        self.tableWidget.setShowGrid(True)
        self.tableWidget.setCornerButtonEnabled(True)
        self.tableWidget.setRowCount(10) ############## 표의 행 갯수
        self.tableWidget.setColumnCount(2) ############ 표의 열 갯수
        self.tableWidget.setObjectName("tableWidget")
        # 카카오톡 ID 헤더 폰트 속성 지정
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setPointSize(13)
        font.setBold(True)
        item.setFont(font)

        self.tableWidget.horizontalHeader().setFixedHeight(50) #헤더 높이 설정
      
        # 밑에 부분 score_1_GUI 배경에 적용 
        #background: QLinearGradient(x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #30E8BF, stop: 1 #ffc850);
        #background: QLinearGradient(x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #2980b9, stop: 1 #8e44ad);
        Score.setStyleSheet(
        '''
        background: QLinearGradient(x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #30E8BF, stop: 1 #FF8235);
   
        '''
        )
        self.tableWidget.setStyleSheet(
        '''
        font-family: 맑은 고딕;
        background: white;
        font-size: 22px;
        background-image: url(images/logo_image.png);
        border: 6px solid gray; border-style: outset;
   
        '''
        )
        # 표 내부 테두리 삭제
        #self.tableWidget.setShowGrid(False)
        # 표 내용 수정 방지
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # 점수 헤더 폰트 속성 지정
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        ##################################3
        item.setBackground(QtCore.Qt.red)
        ################################333
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setPointSize(13)
        font.setBold(True)
        item.setFont(font)
        
        self.tableWidget.setHorizontalHeaderItem(1, item)
        '''
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setFamily("궁서체")
        font.setPointSize(11)
        item.setFont(font)
        '''
        self.tableWidget.setItem(0, 0, item)
        '''
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setFamily("궁서체")
        font.setPointSize(11)
        font.setUnderline(False)
        item.setFont(font)
        '''
        self.tableWidget.setItem(0, 1, item)
        #OK_Button 속성 지정
        self.OK_Button = QtWidgets.QPushButton(Score)
        self.OK_Button.setGeometry(QtCore.QRect(130, 650, 110, 70)) #OK 버튼 위치/크기
        self.OK_Button.setObjectName("OK_Button")
        self.OK_Button.clicked.connect(self.OK_bttnClicked)

        #OK_Button 디자인
        font = QtGui.QFont()
        font.setPointSize(18)
        font.setFamily('맑은 고딕')
        font.setBold(True)
        self.OK_Button.setFont(font)
        self.OK_Button.setStyleSheet(
        '''
        QPushButton { color:white; background:rgb(175, 215, 85);
        border:5px solid rgb(170, 255, 127); border-style:outset;
        border-radius: 25px;}
        QPushButton:hover { color: rgb(59,42,127);}
        '''
        )
        # 마우스 포인터 변경
        self.OK_Button.setCursor(QCursor(QtCore.Qt.PointingHandCursor))

        #Restart_Button 속성 지정
        self.Restart_Button = QtWidgets.QPushButton(Score)
        self.Restart_Button.setGeometry(QtCore.QRect(270, 650, 170, 70))
        self.Restart_Button.setObjectName("Restart_Button")
        self.Restart_Button.clicked.connect(self.Restart_bttnClicked)

        #Restart_Button 디자인
        font = QtGui.QFont()
        font.setPointSize(18)
        font.setFamily('맑은 고딕')
        font.setBold(True)
        self.Restart_Button.setFont(font)
        self.Restart_Button.setStyleSheet(
        '''
        QPushButton { color:white; background:rgb(255, 208, 20);
        border:5px solid rgb(255, 227, 111); border-style:outset;
        border-radius: 25px;}
        QPushButton:hover { color: rgb(59,42,127)}
        '''
        )
        # 마우스 포인터 변경
        self.Restart_Button.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        
        ################ 표 셀 값 지정하기 ################# 
        import openpyxl
        wb = openpyxl.load_workbook('회원관리(테스트용).xlsx')
        ws = wb['회원관리']

        # 카카오톡 ID열 데이터 불어오기
        # for rowNum in range(2, ws.max_row+1):
        for rowNum in range(2, 12):
            kakao_id = str(ws.cell(rowNum, column=1).value) # 카카오톡 ID
            score = str(ws.cell(rowNum, column=2).value) # 점수
            self.tableWidget.setItem(rowNum-2, 0, QTableWidgetItem(kakao_id))
            self.tableWidget.setItem(rowNum-2, 1, QTableWidgetItem(score))

            # 텍스트 가운데 정렬
            self.tableWidget.item(rowNum-2,0).setTextAlignment(QtCore.Qt.AlignCenter)
            self.tableWidget.item(rowNum-2,1).setTextAlignment(QtCore.Qt.AlignCenter)

            # 행 높이 설정
            self.tableWidget.setRowHeight(rowNum-2,55)

        '''
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        '''
        # 열 너비 설정
        self.tableWidget.setColumnWidth(0,300)
        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        ####################################################
        
        self.retranslateUi(Score)
        QtCore.QMetaObject.connectSlotsByName(Score)
        
    def retranslateUi(self, Score):
        _translate = QtCore.QCoreApplication.translate
        Score.setWindowTitle(_translate("Score", "Score"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("Score", "카카오톡 ID"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("Score", "점수"))  
        self.OK_Button.setText(_translate("Score", "Quit"))
        self.Restart_Button.setText(_translate("Score","Restart"))


if __name__ == "__main__":
    import sys
    libpaths = QtWidgets.QApplication.libraryPaths() #추가
    libpaths.append("C:\\Users\사용자\AppData\Local\Programs\Python\Python37-32\Lib\site-packages\PyQt5\Qt\plugins")
    QtWidgets.QApplication.setLibraryPaths(libpaths)
    
    app = QtWidgets.QApplication(sys.argv)
    
    #app.setStyle(QStyleFactory.create('Fusion'))

    Score = QtWidgets.QWidget()
    ui = Ui_Score()
    ui.setupUi(Score)
    Score.show()
    sys.exit(app.exec_())

