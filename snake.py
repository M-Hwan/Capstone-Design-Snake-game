# _*_ coding:utf-8 _*_
# 명환
# 주연 (해상도 변경, 테두리 감지 재설정, 랜덤 아이템 위치 재배열)
# 상현 (사운드)

import cv2
import numpy as np
from time import time
import random
import math
import webcolors
import threading  # 시간(초) 측정
import os  # 파일 리스트 가져오기 위함
import pygame
from random import randint

# 과일 아이템이 들어있는 경로 지정
path_fruits = r"images\fruits"
file_list_fruits = os.listdir(path_fruits)  # 폴더에 있는 모든 파일명을 리스트 형태로 저장
file_dir_list_fruits = []  # 사진별 상대경로를 담을 리스트

# 음식 아이템이 들어있는 경로 지정
path_foods = r"images\food_thing"
file_list_foods = os.listdir(path_foods)  # 폴더에 있는 모든 파일명을 리스트 형태로 저장
file_dir_list_foods = []  # 사진별 상대경로를 담을 리스트

# 감점 아이템이 들어있는 경로 지정
path_bad = r"images\bad_things"
file_list_bad = os.listdir(path_bad)  # 폴더에 있는 모든 파일명을 리스트 형태로 저장
file_dir_list_bad = []  # 사진별 상대경로를 담을 리스트

# 게임종료 아이템이 들어있는 경로 지정
path_end = r"images\end_things"
file_list_end = os.listdir(path_end)  # 폴더에 있는 모든 파일명을 리스트 형태로 저장
file_dir_list_end = []  # 사진별 상대경로를 담을 리스트

# 효과음이 들어있는 경로 지정
path_sound = r"sound"
file_list_sound = os.listdir(path_sound)  # 폴더에 있는 모든 파일명을 리스트 형태로 저장
file_dir_list_sound = []  # 효과음별 상대경로를 담을 리스트

# 과일 아이템 사진별 상대경로 만들기
for i in range(len(file_list_fruits)):
    real_path_fruits = path_fruits + '\\' + file_list_fruits[i]
    file_dir_list_fruits.append(real_path_fruits)

# 음식 아이템 사진별 상대경로 만들기
for i in range(len(file_list_foods)):
    real_path_foods = path_foods + '\\' + file_list_foods[i]
    file_dir_list_foods.append(real_path_foods)

# 감점 아이템 사진별 상대경로 만들기
for i in range(len(file_list_bad)):
    real_path_bad = path_bad + '\\' + file_list_bad[i]
    file_dir_list_bad.append(real_path_bad)

# 게임종료 아이템 사진별 상대경로 만들기
for i in range(len(file_list_end)):
    real_path_end = path_end + '\\' + file_list_end[i]
    file_dir_list_end.append(real_path_end)

# 효과음 음악별 상대경로 만들기

for i in range(len(file_list_sound)):
    real_path_sound = path_sound + '\\' + file_list_sound[i]
    file_dir_list_sound.append(real_path_sound)

'''
print(file_dir_list_sound)
print(file_dir_list_fruits)
print(file_dir_list_foods)
print(file_dir_list_bad)
print(file_dir_list_end)
'''

'''
cv2.imread는 파일경로에 한글과 같은 유니코드가 포함되어 있으면, 인식하지 못함.
따라서,cv2.imdecode를 사용해야 함.
'''
# font = cv2.FONT_HERSHEY_COMPLEX_SMALL
font = cv2.FONT_HERSHEY_DUPLEX


# 랜덤으로 첫번째 아이템 이미지 불러오는 함수
def create_random_item():
    global selected_item, item_list
    # 아이템 구분하는 flag 변수
    # 0 = 50점 추가 / 1 = 100점 추가 / 2 = 50점 감소 / 3 = 100점 감소(shit) / -1 = 게임종료
    # clown = 뱀 꼬리 100 증가 / ghost  = 유령 이미지 커지기 / axe = 점수 1/2 / poison = 테두리 줄이기
    item_flag = 0

    fruits_item = str(random.choice(file_dir_list_fruits))
    foods_item = str(random.choice(file_dir_list_foods))
    bad_item = str(random.choice(file_dir_list_bad))
    end_item = str(random.choice(file_dir_list_end))

    if score >= 0:
        selected_item = random.choice([fruits_item, fruits_item, fruits_item, foods_item, foods_item, bad_item, bad_item, end_item])
    elif score > 300:
        selected_item = random.choice([fruits_item, fruits_item, foods_item, foods_item, bad_item, bad_item, bad_item, end_item])
    elif score > 500:
        selected_item = random.choice([fruits_item, foods_item, bad_item, bad_item, end_item])

    if selected_item == fruits_item:
        item_flag = 0
    elif selected_item == foods_item:
        item_flag = 1
    elif selected_item == bad_item:
        item_flag = 2
    elif selected_item == end_item:
        item_flag = -1

    item = cv2.imdecode(np.fromfile(selected_item, np.uint8), -1)
    item_list = selected_item[18:]
    print(item.shape)
    number_list = item.shape  # 이미지의 3차원 배열을 변수에 저장
    number = int(number_list[2]) - 1
    item_mask = item[:, :, number]
    item_mask_inv = cv2.bitwise_not(item_mask)
    item = item[:, :, 0:3]
    # print('item')
    # print(item)
    # print(item_flag)

    item = cv2.resize(item, (40, 40), interpolation=cv2.INTER_AREA)
    item_mask = cv2.resize(item_mask, (40, 40), interpolation=cv2.INTER_AREA)
    item_mask_inv = cv2.resize(item_mask_inv, (40, 40), interpolation=cv2.INTER_AREA)

    return item, item_mask, item_mask_inv, item_flag

# 랜덤으로 두번째 아이템 이미지 불러오는 함수
def create_random_item1():
    global selected_item1, item_list1
    # 아이템 구분하는 flag 변수
    # 0 = 50점 추가 / 1 = 100점 추가 / 2 = 50점 감소 / 3 = 100점 감소(shit) / -1 = 게임종료
    # clown = 뱀 꼬리 100 증가 / ghost  = 유령 이미지 커지기 / axe = 점수 1/2 / poison = 테두리 줄이기
    item_flag1 = 0

    fruits_item1 = str(random.choice(file_dir_list_fruits))
    foods_item1 = str(random.choice(file_dir_list_foods))
    bad_item1 = str(random.choice(file_dir_list_bad))
    end_item1 = str(random.choice(file_dir_list_end))

    if score >= 0:
        selected_item1 = random.choice([fruits_item1, fruits_item1, fruits_item1, foods_item1, foods_item1, bad_item1, bad_item1, end_item1])
    elif score > 300:
        selected_item1 = random.choice([fruits_item1, fruits_item1, foods_item1, foods_item1, bad_item1, bad_item1, bad_item1, end_item1])
    elif score > 500:
        selected_item1 = random.choice([fruits_item1, foods_item1, bad_item1, bad_item1, end_item1])

    if selected_item1 == fruits_item1:
        item_flag1 = 0
    elif selected_item1 == foods_item1:
        item_flag1 = 1
    elif selected_item1 == bad_item1:
        item_flag1 = 2
    elif selected_item1 == end_item1:
        item_flag1 = -1

    item1 = cv2.imdecode(np.fromfile(selected_item1, np.uint8), -1)
    item_list1 = selected_item1[18:]

    # print(item.shape)
    number_list1 = item1.shape  # 이미지의 3차원 배열을 변수에 저장
    number1 = int(number_list1[2]) - 1
    item_mask1 = item1[:, :, number1]
    item_mask_inv1 = cv2.bitwise_not(item_mask1)
    item1 = item1[:, :, 0:3]
    # print('item')
    # print(item)
    # print(item_flag)

    item1 = cv2.resize(item1, (40, 40), interpolation=cv2.INTER_AREA)
    item_mask1 = cv2.resize(item_mask1, (40, 40), interpolation=cv2.INTER_AREA)
    item_mask_inv1 = cv2.resize(item_mask_inv1, (40, 40), interpolation=cv2.INTER_AREA)

    return item1, item_mask1, item_mask_inv1, item_flag1

# 랜덤으로 두번째 아이템 이미지 불러오는 함수
def create_random_item2():
    global selected_item2, item_list2
    # 아이템 구분하는 flag 변수
    # 0 = 50점 추가 / 1 = 100점 추가 / 2 = 50점 감소 / 3 = 100점 감소(shit) / -1 = 게임종료
    # clown = 뱀 꼬리 100 증가 / ghost  = 유령 이미지 커지기 / axe = 점수 1/2 / poison = 테두리 줄이기
    item_flag2 = 0

    fruits_item2 = str(random.choice(file_dir_list_fruits))
    foods_item2 = str(random.choice(file_dir_list_foods))
    bad_item2 = str(random.choice(file_dir_list_bad))
    end_item2 = str(random.choice(file_dir_list_end))

    if score >= 0:
        selected_item2 = random.choice([fruits_item2, fruits_item2, fruits_item2, foods_item2, foods_item2, bad_item2, bad_item2, end_item2])
    elif score > 300:
        selected_item2 = random.choice([fruits_item2, fruits_item2, foods_item2, foods_item2, bad_item2, bad_item2, bad_item2, end_item2])
    elif score > 500:
        selected_item2 = random.choice([fruits_item2, foods_item2, bad_item2, bad_item2, end_item2])

    if selected_item2 == fruits_item2:
        item_flag1 = 0
    elif selected_item2 == foods_item2:
        item_flag1 = 1
    elif selected_item2 == bad_item2:
        item_flag1 = 2
    elif selected_item2 == end_item2:
        item_flag1 = -1

    item2 = cv2.imdecode(np.fromfile(selected_item2, np.uint8), -1)
    item_list2 = selected_item2[18:]

    # print(item.shape)
    number_list2 = item2.shape  # 이미지의 3차원 배열을 변수에 저장
    number2 = int(number_list2[2]) - 1
    item_mask2 = item2[:, :, number2]
    item_mask_inv2 = cv2.bitwise_not(item_mask2)
    item2 = item2[:, :, 0:3]
    # print('item')
    # print(item)
    # print(item_flag)

    item2 = cv2.resize(item2, (40, 40), interpolation=cv2.INTER_AREA)
    item_mask2 = cv2.resize(item_mask2, (40, 40), interpolation=cv2.INTER_AREA)
    item_mask_inv2 = cv2.resize(item_mask_inv2, (40, 40), interpolation=cv2.INTER_AREA)

    return item2, item_mask2, item_mask_inv2, item_flag2
#########################################################
# (추가 1). 게임 종료 시, DB 데이터 전송 함수 [선 실행]
def save_data(score):
    from openpyxl import load_workbook
    wb = load_workbook("회원관리(테스트용).xlsx")
    ws = wb.active
    wb2 = load_workbook("플레이로그.xlsx")
    ws2 = wb2.active

    if ( score < 0):
        score = 0
    
    ws2.cell(row=ws2.max_row, column=2).value = score
    user_id = ws2.cell(row=ws2.max_row, column=1).value

    ws.cell(row=ws.max_row + 1, column=1).value = user_id
    ws.cell(row=ws.max_row, column=2).value = score

    wb.save("회원관리(테스트용).xlsx")
    wb2.save("플레이로그.xlsx")
    return
#######################################################
# (추가 2). 게임 종료 시, DB 데이터 순위 정렬하여 저장 [후 실행]
def file_ranking():
        import pandas as pd
        from openpyxl import load_workbook

        rank = pd.read_excel('회원관리(테스트용).xlsx')
        rank.sort_values(by='점수', ascending = False).to_excel('회원관리(테스트용).xlsx',index=False)

        wb = load_workbook('회원관리(테스트용).xlsx')
        ws = wb.active
        ws.title = '회원관리'

        for rowNum in range (2, ws.max_row+1):
                rank = int(rowNum-1)
                ws.cell(rowNum, column = 3, value = rank)
            
        ws.column_dimensions['A'].width = 20 # A열의 너비를 20으로 설정
        wb.save('회원관리(테스트용).xlsx')
        return
#######################################################

# 시간 반복용 flag 변수
end = False  # 점수가 0점 이하가 되면, 게임오버 되게 하는 변수
count = 0


# 일정 시간마다 점수 감점 함수
def score_decrease(second=10.0):
    global end, score, count
    if end:
        return None
    if count != 0:
        score -= 10
    count += 1

    # 10초마다 score_decrease 함수 반복실행
    # print('timer 작동됨 1')
    threading.Timer(second, score_decrease, [second]).start()
    # print('timer 작동됨 2')


# 3초마다 아이템을 변경하는 함수
def expired_item(second=3.0):
    global item, item_mask, item_mask_inv, item_flag, random_x, random_y
    if end:
        return None
    # print('random 함수 호출')

    item, item_mask, item_mask_inv, item_flag = create_random_item()
    if item_list == 'poison.png':
        random_x = random.randint(250, 950)
        random_y = random.randint(100, 550)
    else:
        # random_x = random.randint(10, 550)
        # random_y = random.randint(10, 400)
        random_x = random.randint(10, 1050)
        random_y = random.randint(10, 650)
    # print('변경됨')

    # 3초마다 expired_item 함수 반복실행
    # print('timer 작동됨 3')
    threading.Timer(second, expired_item, [second]).start()
    # print('timer 작동됨 4')

def double_item(second=3.0):
    global item1, item_mask1, item_mask_inv1, item_flag1, random_x1, random_y1
    if end:
        return None
    # print('random 함수 호출')
    item1, item_mask1, item_mask_inv1, item_flag1 = create_random_item1()
    random_x1 = random.randint(10, 550)
    random_y1 = random.randint(10, 400)
    # print('변경됨')

    # 3초마다 expired_item 함수 반복실행
    # print('timer 작동됨 3')
    threading.Timer(second, double_item, [second]).start()
    # print('timer 작동됨 4')

def triple_item(second=3.0):
    global item2, item_mask2, item_mask_inv2, item_flag2, random_x2, random_y2
    if end:
        return None
    # print('random 함수 호출')
    item2, item_mask2, item_mask_inv2, item_flag2 = create_random_item2()
    random_x2 = random.randint(10, 550)
    random_y2 = random.randint(10, 400)
    # print('변경됨')

    # 3초마다 expired_item 함수 반복실행
    # print('timer 작동됨 3')
    threading.Timer(second, triple_item, [second]).start()
    # print('timer 작동됨 4')


###################################################
# blank_img = np.zeros((480, 640, 3), np.uint8)
blank_img = np.zeros((720, 1280, 3), np.uint8)

video = cv2.VideoCapture(0, cv2.CAP_DSHOW)
# 비디오 사이즈 고정용
# video.set(3, 640)
# video.set(4, 480)

video.set(3, 1280)
video.set(4, 720)

# 배경음악 초기화
pygame.mixer.init()
pygame.mixer.pre_init(44100, -16, 2, 512)

# 게임 효과음
shit_sound = pygame.mixer.Sound('sound\\shit.wav')
clown_sound = pygame.mixer.Sound('sound\\clown.wav')
skeleton_sound = pygame.mixer.Sound('sound\\skeleton.wav')
bomb_sound = pygame.mixer.Sound('sound\\bomb.wav')
clover_sound = pygame.mixer.Sound('sound\\clover.wav')
double_sound = pygame.mixer.Sound('sound\\double.wav')
ghost_sound = pygame.mixer.Sound('sound\\ghost.wav')
poison_sound = pygame.mixer.Sound('sound\\poison.wav')
axe_sound = pygame.mixer.Sound('sound\\axe.wav')
good_sound1 = pygame.mixer.Sound('sound\\sound1.wav')
good_sound2 = pygame.mixer.Sound('sound\\sound2.wav')
good_sound3 = pygame.mixer.Sound('sound\\sound3.wav')
good_sound4 = pygame.mixer.Sound('sound\\sound4.wav')
gameover_sound = pygame.mixer.Sound('sound\\game_over.wav')

# 좋은 효과음 세트
good_sound = {'1': good_sound1, '2': good_sound2, '3': good_sound3, '4': good_sound4}

# 배경음악 실행
pygame.mixer.music.load('sound\\background_sound.wav')
pygame.mixer.music.set_volume(0.6)
pygame.mixer.music.play()

kernel_erode = np.ones((4, 4), np.uint8)
kernel_close = np.ones((15, 15), np.uint8)
# 딱풀 색 설정
# color = 'yellow'
rgb = webcolors.name_to_rgb(u'yellow')
red = rgb.red
blue = rgb.blue
green = rgb.green
# red = random.randint(0,255)
# blue = random.randint(0,255)
# green = random.randint(0,255)

lower_upper = []


def color_convert(r, bl, g):
    # 색상 값을 이진수 데이터로 변환
    co = np.uint8([[[bl, g, r]]])
    hsv_color = cv2.cvtColor(co, cv2.COLOR_BGR2HSV)

    hue = hsv_color[0][0][0]
    lower_upper.append([hue - 10, 100, 100])
    lower_upper.append([hue + 10, 255, 255])
    return lower_upper


def detect_color(h):
    lu = color_convert(red, blue, green)
    # 노란색 색상의 인식 범위 지정(numpy 배열 형식)
    lower = np.array(lu[0])
    upper = np.array(lu[1])
    mask = cv2.inRange(h, lower, upper)
    # Erosion 알고리즘 사용
    # 흰색 오브젝트의 외곽 픽셀을 0(검은색)
    # 노이즈(작은 흰색 물체)를 제거하거나 붙어 있는 오브젝트들을 분리
    mask = cv2.erode(mask, kernel_erode, iterations=1)
    # 이미지 변환 기법(Closing 사용 - 노이즈 제거하여 보다 선명한 이미지로 만듦.)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel_close)
    return mask


def orientation(p, q, r):
    val = int(((q[1] - p[1]) * (r[0] - q[0])) - ((q[0] - p[0]) * (r[1] - q[1])))
    if val == 0:
        return 0
    elif val > 0:
        return 1
    else:
        return 2


def intersect(p, q, r, s):
    o1 = orientation(p, q, r)
    o2 = orientation(p, q, s)
    o3 = orientation(r, s, p)
    o4 = orientation(r, s, q)
    if o1 != o2 and o3 != o4:
        return True

    return False


##############################################
start_time = int(time())
q, snake_len, score, temp = 0, 200, 0, 1
# q, snake_len, score, temp = 0, 200, 500, 1
point_x, point_y = 0, 0
last_point_x, last_point_y, dist, length = 0, 0, 0, 0
points = []
list_len = []

random_x = random.randint(40, 1100)
random_y = random.randint(30, 640)
random_x1 = random.randint(40, 1100)
random_y1 = random.randint(30, 640)
random_x2 = random.randint(40, 1100)
random_y2 = random.randint(30, 640)

a, b, c, d = [], [], [], []
################# 추가 2 ###############
item, item_mask, item_mask_inv, item_flag = create_random_item()
item1, item_mask1, item_mask_inv1, item_flag1 = create_random_item1()
item2, item_mask2, item_mask_inv2, item_flag2 = create_random_item2()


#########################################

# 10초마다 점수 감소하도록 조절
score_decrease()
# 3초마다 아이템 변경되도록 조절
expired_item()
double_item()
triple_item()

# 화면 가리는 유령 이미지 read 및 mask 생성
disturb = cv2.imdecode(np.fromfile(r'images\ghost_opacity.png', np.uint8), -1)
disturb_list = disturb.shape  # 이미지의 3차원 배열을 변수에 저장
disturb_number = int(disturb_list[2]) - 1
disturb_mask = disturb[:, :, disturb_number]
disturb_mask_inv = cv2.bitwise_not(disturb_mask)
disturb = disturb[:, :, 0:3]
disturb = cv2.resize(disturb, (700, 700), interpolation=cv2.INTER_AREA)
disturb_mask = cv2.resize(disturb_mask, (700, 700), interpolation=cv2.INTER_AREA)
disturb_mask_inv = cv2.resize(disturb_mask_inv, (700, 700), interpolation=cv2.INTER_AREA)

# ghost, poison 시간변수 설정
# ghost_time = None
# poison_time = None
bool_ghost = False
bool_poison = False

while True:
    xr, yr, wr, hr = 0, 0, 0, 0
    ret, frame = video.read()
    # 영상 좌우반전
    frame = cv2.flip(frame, 1)
    
    if q == 0 and point_x != 0 and point_y != 0:
        last_point_x = point_x
        last_point_y = point_y
        q = 1

    # 검출된 색상을 HSV 변환
    '''
    Video로 부터 Frame을 읽어 들입니다.
    frame을 HSV로 변환을 합니다.
    변환한 이미지에서 yellow 영역을 찾아서 mask를 생성합니다.
    '''
    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

    # 색상(노란색) 감지 함수를 호출하고, 해당 마스크 값을 mask 변수에 저장
    mask = detect_color(hsv)
    
    # finding contours
    contour, _ = cv2.findContours(mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    # drawing rectangle around the accepted blob
    try:
        for i in range(0, 10):
            xr, yr, wr, hr = cv2.boundingRect(contour[i])
            if (wr * hr) > 2000:
                break
    except:
        pass

    # print('xr: %d, yr: %d, wr: %d, hr: %d' % (xr, yr, wr, hr))

    # 플레이 오브젝트 인식선 그리기
    cv2.rectangle(frame, (xr, yr), (xr + wr, yr + hr), (0, 0, 255), 2)
    # 플레이 화면에 테두리 그리기
    # cv2.rectangle(frame, (0, 0), (640, 480), (255, 0, 0), 10)

    # 1. 명환님 컴퓨터 환경에 맞춘 테두리
    cv2.rectangle(frame, (0, 0), (1280, 720), (255, 0, 0), 20)

    # 2. 주연 컴퓨터 환경에 맞춘 테두리
    # cv2.rectangle(frame, (0,0), (1275, 690), (255, 0, 0), 25)

    point_x = int(xr + (wr / 2))
    point_y = int(yr + (hr / 2))

    dist = int(math.sqrt(pow((last_point_x - point_x), 2) + pow((last_point_y - point_y), 2)))
    ##########
    # print(dist)
    if point_x != 0 and point_y != 0 and dist > 5:
        list_len.append(dist)
        length += dist
        last_point_x = point_x
        last_point_y = point_y
        points.append([point_x, point_y])

    if length >= snake_len:
        for i in range(len(list_len)):
            length -= list_len[0]
            list_len.pop(0)
            points.pop(0)
            if length <= snake_len:
                break

    # blank_img = np.zeros((480, 640, 3), np.uint8)
    blank_img = np.zeros((720, 1280, 3), np.uint8)

    for i, j in enumerate(points):
        if i == 0:
            continue
        # 라인 그리는 부분
        #cv2.line(blank_img, (points[i - 1][0], points[i - 1][1]), (j[0], j[1]), (blue, green, red), 5)
        #cv2.line(blank_img, (points[i - 1][0], points[i - 1][1]), (j[0], j[1]), (255,0,255), 5)
        
        # 뱀 몸통 색상 무지개로 표현
        b = random.randint(0,255)
        g = random.randint(0,255)
        r = random.randint(0,255)
        cv2.line(blank_img, (points[i - 1][0], points[i - 1][1]), (j[0], j[1]), (b,g,r), 5)
    
    cv2.circle(blank_img, (last_point_x, last_point_y), 5, (10, 200, 150), -1)
    ################################################################
    # 뱀 머리 위치 좌표: last_point_x, last_point_y
    # 아이템 좌표: random_x, random_y
    # 아이템 사이즈: random_x+40, random_y+40
    if random_x < last_point_x < (random_x + 40) and random_y < last_point_y < (random_y + 40):
        # 1~4 랜덤 난수 생성(효과음 랜덤 재생시 필요)
        ab = randint(1, 4)
        ab = str(ab)

        # 아이템 구분하는 flag 변수
        # 0 = 50점 추가 / 1 = 100점 추가 / 2 = 50점 감소 / 3 = 100점 감소(shit) / -1 = 게임종료
        # clown = 뱀 꼬리 100 증가 / ghost  = 유령 이미지 커지기 / axe = 점수 1/2 / poison = 테두리 줄이기
        if item_flag == 0:
            good_sound_list = good_sound[ab]
            good_sound_list.play()
            score += 50
        elif item_flag == 1:
            if item_list == 'clover.png':
                clover_sound.play()
                snake_len -= 100  # 점수는 올려주지 않고, 대신 꼬리만 줄여주기
            elif item_list == 'double point.png':
                double_sound.play()
                score *= 2
            else:
                good_sound_list = good_sound[ab]
                good_sound_list.play()
                score += 100
        elif item_flag == 2:
            if item_list == 'ghost.png':
                bool_ghost = True
                ghost_sound.play()
                ghost_time = int(time())
            elif item_list == 'shit.png':
                shit_sound.play()
                score -= 100
            elif item_list == 'axe.png':
                axe_sound.play()
                score //= 2
            elif item_list == 'clown.png':
                clown_sound.play()
                snake_len += 100
            elif item_list == 'poison.png':
                bool_poison = True
                poison_sound.play()
                poison_time = int(time())
            else:
                score -= 50
        elif item_flag == -1:
            if item_list == 'bomb.png':
                bomb_sound.play()
            elif item_list == 'skeleton.png':
                skeleton_sound.play()
            break

        '''
        random_x = random.randint(10, 550)
        random_y = random.randint(10, 400)
        '''
        random_x = random.randint(40, 1100)
        random_y = random.randint(30, 640)

        ######################## 추가 3 ############
        item, item_mask, item_mask_inv, item_flag = create_random_item()  # 점수가 오를때마다,랜덤 아이템 생성 함수 호출
        # 아이템을 먹을 시, 뱀 몸통이 길어짐
        snake_len += 40

        ############################################
    elif random_x1 < last_point_x < (random_x1 + 40) and random_y1 < last_point_y < (random_y1 + 40):
        # 1~4 랜덤 난수 생성(효과음 랜덤 재생시 필요)
        ab = randint(1, 4)
        ab = str(ab)

        # 아이템 구분하는 flag 변수
        # 0 = 50점 추가 / 1 = 100점 추가 / 2 = 50점 감소 / 3 = 100점 감소(shit) / -1 = 게임종료
        # clown = 뱀 꼬리 100 증가 / ghost  = 유령 이미지 커지기 / axe = 점수 1/2 / poison = 테두리 줄이기
        if item_flag1 == 0:
            good_sound_list = good_sound[ab]
            good_sound_list.play()
            score += 50
        elif item_flag1 == 1:
            if item_list1 == 'clover.png':
                clover_sound.play()
                snake_len -= 100  # 점수는 올려주지 않고, 대신 꼬리만 줄여주기
            elif item_list1 == 'double point.png':
                double_sound.play()
                score *= 2
            else:
                good_sound_list = good_sound[ab]
                good_sound_list.play()
                score += 100
        elif item_flag1 == 2:
            if item_list1 == 'ghost.png':
                bool_ghost = True
                ghost_sound.play()
                ghost_time = int(time())
            elif item_list1 == 'shit.png':
                shit_sound.play()
                score -= 100
            elif item_list1 == 'axe.png':
                axe_sound.play()
                score //= 2
            elif item_list1 == 'clown.png':
                clown_sound.play()
                snake_len += 100
            elif item_list1 == 'poison.png':
                bool_poison = True
                poison_sound.play()
                poison_time = int(time())
            else:
                score -= 50
        elif item_flag1 == -1:
            if item_list1 == 'bomb.png':
                bomb_sound.play()
            elif item_list1 == 'skeleton.png':
                skeleton_sound.play()
            break

        '''
        random_x = random.randint(10, 550)
        random_y = random.randint(10, 400)
        '''
        random_x1 = random.randint(40, 1100)
        random_y1 = random.randint(30, 640)

        ######################## 추가 3 ############
        item1, item_mask1, item_mask_inv1, item_flag1 = create_random_item1()  # 점수가 오를때마다,랜덤 아이템 생성 함수 호출
        # 아이템을 먹을 시, 뱀 몸통이 길어짐
        snake_len += 40

        ############################################
    elif random_x2 < last_point_x < (random_x2 + 40) and random_y2 < last_point_y < (random_y2 + 40):
        # 1~4 랜덤 난수 생성(효과음 랜덤 재생시 필요)
        ab = randint(1, 4)
        ab = str(ab)

        # 아이템 구분하는 flag 변수
        # 0 = 50점 추가 / 1 = 100점 추가 / 2 = 50점 감소 / 3 = 100점 감소(shit) / -1 = 게임종료
        # clown = 뱀 꼬리 100 증가 / ghost  = 유령 이미지 커지기 / axe = 점수 1/2 / poison = 테두리 줄이기
        if item_flag2 == 0:
            good_sound_list = good_sound[ab]
            good_sound_list.play()
            score += 50
        elif item_flag2 == 1:
            if item_list2 == 'clover.png':
                clover_sound.play()
                snake_len -= 100  # 점수는 올려주지 않고, 대신 꼬리만 줄여주기
            elif item_list2 == 'double point.png':
                double_sound.play()
                score *= 2
            else:
                good_sound_list = good_sound[ab]
                good_sound_list.play()
                score += 100
        elif item_flag2 == 2:
            if item_list2 == 'ghost.png':
                bool_ghost = True
                ghost_sound.play()
                ghost_time = int(time())
            elif item_list2 == 'shit.png':
                shit_sound.play()
                score -= 100
            elif item_list2 == 'axe.png':
                axe_sound.play()
                score //= 2
            elif item_list2 == 'clown.png':
                clown_sound.play()
                snake_len += 100
            elif item_list2 == 'poison.png':
                bool_poison = True
                poison_sound.play()
                poison_time = int(time())
            else:
                score -= 50
        elif item_flag2 == -1:
            if item_list2 == 'bomb.png':
                bomb_sound.play()
            elif item_list2 == 'skeleton.png':
                skeleton_sound.play()
            break

        '''
        random_x = random.randint(10, 550)
        random_y = random.randint(10, 400)
        '''
        random_x2 = random.randint(40, 1100)
        random_y2 = random.randint(30, 640)

        ######################## 추가 3 ############
        item2, item_mask2, item_mask_inv2, item_flag2 = create_random_item2()  # 점수가 오를때마다,랜덤 아이템 생성 함수 호출
        # 아이템을 먹을 시, 뱀 몸통이 길어짐
        snake_len += 40

        ############################################
    frame = cv2.add(frame, blank_img)
    '''
    roi = frame[random_y:random_y+40, random_x:random_x+40]
    img_bg = cv2.bitwise_and(roi, roi, mask=apple_mask_inv)
    img_fg = cv2.bitwise_and(apple, apple, mask=apple_mask)
    dst = cv2.add(img_bg, img_fg)
    frame[random_y:random_y + 40, random_x:random_x + 40] = dst
    cv2.putText(frame, str("Score - "+str(score)), (250, 450), font, 1, (0, 0, 0), 2, cv2.LINE_AA)
    
    ################추가 부분
    roi_2 = frame[random_y:random_y+40, random_x:random_x+40]
    img_bg2 = cv2.bitwise_and(roi_2, roi_2, mask=lemon_mask_inv)
    img_fg2 = cv2.bitwise_and(lemon, lemon, mask=lemon_mask)
    dst_2 = cv2.add(img_bg2, img_fg2)
    frame[random_y:random_y + 40, random_x:random_x + 40] = dst_2
    #cv2.putText(frame, str("Score - "+str(score)), (250, 450), font, 1, (0, 0, 0), 2, cv2.LINE_AA)
    #########################
    '''
    ############### 추가 4 ##################
    if bool_ghost == True:
        print('this is ROI ghost')
        roi = frame[10:710, 250:950]
        img_bg = cv2.bitwise_and(roi, roi, mask=disturb_mask_inv)
        img_fg = cv2.bitwise_and(disturb, disturb, mask=disturb_mask)
        dst = cv2.add(img_bg, img_fg)
        frame[10:710, 250:950] = dst
        print(time() - ghost_time)
        if (int(time()) - ghost_time) > 3:
            bool_ghost = False
    elif bool_poison == True:
        print('this is ROI poison')
        cv2.rectangle(frame, (220, 80), (1060, 640), (255, 0, 0), 10)
        print(time() - poison_time)
        # 좁은 사각형 벗어났을 때 게임 끝나도록 설정

        # 가로 = 840, 세로 = 560
        random_x = random.randint(250, 950)
        random_y = random.randint(100, 550)

        if xr >= (1060 - (wr / 2) ) or xr <= (220 - (wr / 2) ):
            print('poison x축')
            break
        if yr <= (80 - (hr / 2)) or yr >= (640 - (hr / 2) ):
            print('poison y축')
            break

        if (int(time()) - poison_time) > 3:
            bool_poison = False
    else:
        ############### 첫번째 아이템 ##################
        if score >= 0:
            roi = frame[random_y:random_y + 40, random_x:random_x + 40]
            img_bg = cv2.bitwise_and(roi, roi, mask=item_mask_inv)
            img_fg = cv2.bitwise_and(item, item, mask=item_mask)
            dst = cv2.add(img_bg, img_fg)
            frame[random_y:random_y + 40, random_x:random_x + 40] = dst

        ############### 두번째 아이템 ##################
        if score > 300:
            roi1 = frame[random_y1:random_y1 + 40, random_x1:random_x1 + 40]
            img_bg1 = cv2.bitwise_and(roi1, roi1, mask=item_mask_inv1)
            img_fg1 = cv2.bitwise_and(item1, item1, mask=item_mask1)
            dst1 = cv2.add(img_bg1, img_fg1)
            frame[random_y1:random_y1 + 40, random_x1:random_x1 + 40] = dst1

        ############### 세번째 아이템 ##################
        if score > 500:
            roi2 = frame[random_y2:random_y2 + 40, random_x2:random_x2 + 40]
            img_bg2 = cv2.bitwise_and(roi2, roi2, mask=item_mask_inv2)
            img_fg2 = cv2.bitwise_and(item2, item2, mask=item_mask2)
            dst2 = cv2.add(img_bg2, img_fg2)
            frame[random_y2:random_y2 + 40, random_x2:random_x2 + 40] = dst2

    # 1. 명환님 컴퓨터 환경에 맞춘 총점 표시 위치
    # cv2.putText(frame, str("Score : " + str(score)), (550, 700), font, 1, (0, 0, 0), 2, cv2.LINE_AA)
    # 2. 주연 컴퓨터 환경에 맞춘 총점 표시 위치
    cv2.putText(frame, str("Score : " + str(score)), (550, 660), font, 1, (0, 0, 200), 2, cv2.LINE_AA)
    ##########################################
    # 플레이시간 표시
    # play_time = 'play time : ' + str(int(time())-start_time)
    play_time = int(time() - start_time)

    # 1. 명환님 컴퓨터 환경에 맞춘 총 플레이 시간 표시 위치
    # cv2.putText(frame, str('play time : ' + str(play_time)), (950, 50), font, 1, (0, 0, 0), 2, cv2.LINE_AA)

    # 2. 주연 컴퓨터 환경에 맞춘 총 플레이 시간 표시 위치
    cv2.putText(frame, str('play time : ' + str(play_time) + ' sec'), (950, 60), font, 1, (0, 0, 0), 2, cv2.LINE_AA)

    if len(points) > 5:
        b = points[len(points) - 2]
        a = points[len(points) - 1]
        for i in range(len(points) - 3):
            c = points[i]
            d = points[i + 1]
            if intersect(a, b, c, d) and len(c) != 0 and len(d) != 0:
                temp = 0
                break
        if temp == 0:
            break

    ## 아이템 띄우는 모듈!!!##############3
    cv2.imshow("frame", frame)
    #####################################3
    # 시간마다 뱀 몸통 길이가 길어지는 것이 아니라, 아이템을 먹을 떄마다 길어지게 구현!
    '''
    if (int(time()) - start_time) > 1:
        snake_len += 40
        start_time = int(time())
    '''
    key = cv2.waitKey(1)
    if key == 27:
        break
    if score < 0:
        end = True
        break

    # 테두리에 닿으면 게임오버
    # xr, yr = x, y 좌표 / wr, hr = 사각형의 너비, 높이
    if play_time >= 10:
        if xr == 0 and yr == 0 and wr == 0 and hr == 0:
            print('화면 밖')
            break

        if xr >= 1085:
            print('오른쪽 테두리')
            break

        if (xr <= 165 and wr <= 35):
            print('왼쪽 테두리')
            break

        if (yr <= 20 and hr <= 45):
            print('위쪽 테두리')
            break

        if yr >= 615:
            print('아래 테두리')
            break

# 배경음악 중지
pygame.mixer.music.stop()
# 게임 오버
gameover_sound.set_volume(1)
gameover_sound.play()

############################################################################
# Threading 종료용 변수
end = True
###########################################################################
# 게임 종료 시, 데이터 DB에 저장하는 함수 호출
save_data(score)
############################################
# 게임 종료 시, DB에 저장된 점수 순위 정렬 함수 호출
file_ranking()
############################################

video.release()
# cv2.destroyAllWindows()
# 색상은 (B, G, R)로 표현!
cv2.putText(frame, str("Game Over!"), (350, 300), font, 3, (0, 0, 255), 3, cv2.LINE_AA)
cv2.putText(frame, str("Press any key to Exit"), (450, 380), font, 1, (0, 228, 255), 2, cv2.LINE_AA)
cv2.imshow("frame", frame)
k = cv2.waitKey(0)
f = open('newCorpus_GUI.py', 'rt', encoding='UTF8')
exec(f.read())
f.close()
print('k:', k)
cv2.destroyAllWindows()
