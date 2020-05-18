# -*- coding: utf-8 -*-

# Score implementation generated from reading ui file 'score.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!


# 이주연
# 게임 시작 화면, 점수판 GUI

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *

class Ui_Score(object):
    
    def OK_bttnClicked(self):
        #sys.exit(0)
        Score.hide()
        
    def setupUi(self, Score):
        
        Score.setObjectName("Score")
        Score.resize(350, 600)    ##### 창 크기:(가로,세로)
        self.tableWidget = QtWidgets.QTableWidget(Score)
        # 위젯 위치 조절(가로로 떨어진 거리, 세로로 떨어진 거리, 가로, 세로)
        self.tableWidget.setGeometry(QtCore.QRect(10, 10, 335, 488)) 
        self.tableWidget.setShowGrid(True)
        self.tableWidget.setCornerButtonEnabled(True)
        self.tableWidget.setRowCount(10) ############## 표의 행 갯수
        self.tableWidget.setColumnCount(2) ############ 표의 열 갯수
        self.tableWidget.setObjectName("tableWidget")
        # 카카오톡 ID 칸 폰트 속성 지정
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        # 점수 칸 폰트 속성 지정
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        
        self.tableWidget.setHorizontalHeaderItem(1, item)
        '''
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setFamily("a옛날목욕탕B")
        font.setPointSize(11)
        item.setFont(font)
        '''
        self.tableWidget.setItem(0, 0, item)
        '''
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setFamily("a옛날목욕탕B")
        font.setPointSize(11)
        font.setUnderline(False)
        item.setFont(font)
        '''
        self.tableWidget.setItem(0, 1, item)
        
        #OK_Button 속성 지정
        self.OK_Button = QtWidgets.QPushButton(Score)
        self.OK_Button.setGeometry(QtCore.QRect(130, 520, 90, 60)) #OK 버튼 위치/크기
        self.OK_Button.setObjectName("OK_Button")
        self.OK_Button.clicked.connect(self.OK_bttnClicked) # 이벤트 핸들러
        
        #OK_Button 디자인
        font = QtGui.QFont()
        font.setPointSize(17)
        font.setFamily('맑은 고딕')
        font.setBold(True)
        self.OK_Button.setFont(font)
        self.OK_Button.setStyleSheet(
       '''
        QPushButton { color:white; background:rgb(175, 215, 85);
        border:5px solid rgb(170, 255, 127); border-style:outset}
        QPushButton:hover { color: rgb(59,42,127)}
        '''
        )
        
        ################ 표 셀 값 지정하기 ################# 
        import openpyxl
        wb = openpyxl.load_workbook('회원관리(테스트용).xlsx')
        ws = wb['회원관리']

        # 카카오톡 ID열 데이터 불어오기
        for rowNum in range(2, ws.max_row+1):
            kakao_id = str(ws.cell(rowNum, column=1).value) # 카카오톡 ID
            score = str(ws.cell(rowNum, column=2).value) # 점수
            self.tableWidget.setItem(rowNum-2, 0, QTableWidgetItem(kakao_id))
            self.tableWidget.setItem(rowNum-2, 1, QTableWidgetItem(score))
        ####################################################
        
        self.retranslateUi(Score)
        QtCore.QMetaObject.connectSlotsByName(Score)
        
    def retranslateUi(self, Score):
        _translate = QtCore.QCoreApplication.translate
        Score.setWindowTitle(_translate("Score", "Score"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("Score", "카카오톡 ID"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("Score", "점수"))  
        self.OK_Button.setText(_translate("Score", "OK"))
        
if __name__ == "__main__":
    import sys
    libpaths = QtWidgets.QApplication.libraryPaths() #추가
    libpaths.append("C:\\Users\사용자\AppData\Local\Programs\Python\Python37-32\Lib\site-packages\PyQt5\Qt\plugins")
    QtWidgets.QApplication.setLibraryPaths(libpaths)
   
    app = QtWidgets.QApplication(sys.argv)
    Score = QtWidgets.QWidget()
    ui = Ui_Score()
    ui.setupUi(Score)
    Score.show()
    sys.exit(app.exec_())

