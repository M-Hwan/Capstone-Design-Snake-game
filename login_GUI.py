# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'login_GUI.ui'
#
# Created by: PyQt5 UI code generator 5.13.0
#
# WARNING! All changes made in this file will be lost!

# 창 연결, DB 연동, 디자인 추가(주연)

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from game_rules import *

class Ui_Login(object):

    def pushButton(self,title,content):
        msgBox = QMessageBox.about(None, title, content)
    
    def show_game_method(self):
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_Rules()
        self.ui.setupUi(self.window)
        #Login.close()
        #Login.hide() # 로그인 화면 숨기기
        self.window.show()
        
    def OK_bttnClicked(self):
        user_id = self.id_lineEdit.text()

        # 1. 아이디를 입력하지 않은 경우
        if (user_id == ''):
            warning = '아이디를 입력해주세요.'
            self.pushButton('로그인 실패',warning)
            return
          
        # 2. 아이디를 입력한 경우
        from openpyxl import load_workbook
        wb = load_workbook("회원관리(테스트용).xlsx")
        ws = wb.active
        wb1 = load_workbook("플레이로그.xlsx")
        ws1 = wb1.active

        for rowNum in range(2, ws.max_row+1):
            # (2-1). 입력한 아이디가 DB에 존재하는 경우 (로그인 성공)
            if (user_id == str(ws.cell(row=rowNum, column = 1).value)):
                content = '로그인에 성공했습니다.\n게임을 시작합니다.\n'
                self.pushButton('로그인 성공',content)

                r= str(ws1.max_row+1)
                ws1['A'+r] = user_id
                wb1.save("플레이로그.xlsx")
                self.show_game_method() # 플레이 방법 GUI 표시 함수 호출
                return
            
        # (2-2). 입력한 아이디가 DB에 존재하지 않는 경우 (로그인 실패)
        content2 = 'ID를 재입력하거나,\n카카오톡 플러스 친구를 먼저 등록해주세요.'       
        self.pushButton('로그인 실패',content2)
        return
        
    def setupUi(self, Login):
        Login.setObjectName("Login")
        Login.resize(889, 770)
  
        # start_button 속성 지정
        self.OK_button = QtWidgets.QPushButton(Login)
        self.OK_button.setGeometry(QtCore.QRect(370, 690, 141, 60))
        self.OK_button.setObjectName("OK_button")
        font = QtGui.QFont()
        font.setPointSize(17)
        font.setFamily('맑은 고딕')
        font.setBold(True)
        self.OK_button.setFont(font)
        #1. 보라
        #self.OK_button.setStyleSheet('color:white;background:rgb(170, 170, 255);border: 5px solid rgb(212, 214, 255);border-style:outset')
        #2. 초록
        #self.OK_button.setStyleSheet('color:white;background:rgb(85, 170, 0);border: 5px solid rgb(170, 255, 127);border-style:outset')
        #self.OK_button.setStyleSheet('color:white;background:rgb(175, 215, 85);border: 5px solid rgb(170, 255, 127);border-style:outset')
        self.OK_button.setStyleSheet(
        '''
        QPushButton { color:white; background:rgb(175, 215, 85);
        border:5px solid rgb(170, 255, 127); border-style:outset}
        QPushButton:hover { color: rgb(59,42,127)}
        '''
        )
        #3. 노랑
        
        #self.OK_button.setStyleSheet(
        #'''
        #QPushButton { color:white; background:rgb(255, 208, 20);
        #border:5px solid rgb(255, 227, 111); border-style:outset}
        #QPushButton:hover { color: rgb(59,42,127)}
        #'''
        #)
        
        #4. 파랑
        #self.OK_button.setStyleSheet('color:white;background:rgb(59, 42, 127);border: 5px solid rgb(59,42,217);border-style:outset')
        
        self.OK_button.clicked.connect(self.OK_bttnClicked) #이벤트 핸들러

        # id_lineEdit 속성 지정
        self.id_lineEdit = QLineEdit(Login)
        self.id_lineEdit.setGeometry(QtCore.QRect(420, 620, 341, 51))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.id_lineEdit.setFont(font)
        self.id_lineEdit.setObjectName("id_lineEdit")
        
        # label_1 속성 지정
        self.label = QtWidgets.QLabel(Login)
        self.label.setGeometry(QtCore.QRect(2, 0, 891, 601))
        self.label.setText("")
        self.label.setPixmap(QtGui.QPixmap("images/notice.png"))
        self.label.setObjectName("label")
        
        # label_2 속성 지정
        self.label_2 = QtWidgets.QLabel(Login)
        self.label_2.setGeometry(QtCore.QRect(160, 620, 241, 55))
        self.label_2.setText("")
        self.label_2.setPixmap(QtGui.QPixmap("images/ID입력.png"))
        self.label_2.setObjectName("label_2")
        
        self.retranslateUi(Login)
        QtCore.QMetaObject.connectSlotsByName(Login)

    def retranslateUi(self, Login):
        _translate = QtCore.QCoreApplication.translate
        Login.setWindowTitle(_translate("Login", "Login"))
        self.OK_button.setText(_translate("Login", "확   인"))
        #self.id_label.setText(_translate("Login", "아이디"))

if __name__ == "__main__":
    import sys
    libpaths = QtWidgets.QApplication.libraryPaths() #추가
    libpaths.append("C:\\Users\사용자\AppData\Local\Programs\Python\Python37-32\Lib\site-packages\PyQt5\Qt\plugins")
    QtWidgets.QApplication.setLibraryPaths(libpaths)
    
    app = QtWidgets.QApplication(sys.argv)
    Login = QtWidgets.QDialog()
    ui = Ui_Login()
    ui.setupUi(Login)
    Login.show()
    sys.exit(app.exec_())
